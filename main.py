import io
import re
from fastapi import FastAPI, File, UploadFile, HTTPException
from fastapi.responses import StreamingResponse, HTMLResponse
import pdfplumber
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

app = FastAPI(title="Professional PDF to Master Dashboard Converter")

# -------------------------------------------------------------
# 🎨 WEB UI INTERFACE
# -------------------------------------------------------------
HTML_PAGE = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>PDF to Executive Master Sheet Converter</title>
    <style>
        * { box-sizing: border-box; margin: 0; padding: 0; }
        body { font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; background: #eef2f5; color: #333; display: flex; justify-content: center; align-items: center; min-height: 100vh; padding: 20px; }
        .card { background: #ffffff; width: 100%; max-width: 500px; padding: 32px; border-radius: 12px; box-shadow: 0 4px 20px rgba(0,0,0,0.08); text-align: center; }
        h2 { color: #1B365D; margin-bottom: 8px; font-size: 22px; font-weight: 700; }
        p { color: #666; font-size: 13px; margin-bottom: 24px; }
        .upload-area { border: 2px dashed #007bff; border-radius: 8px; padding: 30px 20px; background: #f8faff; cursor: pointer; transition: 0.3s; margin-bottom: 20px; }
        .upload-area:hover { background: #eaf2ff; }
        .file-input { display: none; }
        .icon { font-size: 40px; color: #007bff; margin-bottom: 10px; }
        .btn { width: 100%; padding: 12px; background: #1B365D; color: white; border: none; border-radius: 6px; font-size: 15px; font-weight: bold; cursor: pointer; transition: 0.3s; }
        .btn:hover { background: #122440; }
        .btn:disabled { background: #ccc; cursor: not-allowed; }
        #file-name { margin-top: 10px; font-size: 12px; font-weight: bold; color: #007bff; }
        .spinner { display: none; margin: 15px auto 0; border: 4px solid #f3f3f3; border-top: 4px solid #1B365D; border-radius: 50%; width: 30px; height: 30px; animation: spin 1s linear infinite; }
        @keyframes spin { 0% { transform: rotate(0deg); } 100% { transform: rotate(360deg); } }
    </style>
</head>
<body>
    <div class="card">
        <h2>📄 Executive Master Sheet</h2>
        <p>PDF Upload Karein & Designer Pro Excel Dashboard Download Karein</p>
        <form id="uploadForm">
            <div class="upload-area" onclick="document.getElementById('pdfFile').click()">
                <div class="icon">📁</div>
                <div style="font-weight: bold; font-size: 14px;">Select PDF File</div>
                <div id="file-name">No file selected</div>
            </div>
            <input type="file" id="pdfFile" class="file-input" accept=".pdf" onchange="showFileName()" required>
            <button type="submit" id="submitBtn" class="btn">⚡ Generate Designer Excel</button>
        </form>
        <div id="spinner" class="spinner"></div>
    </div>

    <script>
        function showFileName() {
            const input = document.getElementById('pdfFile');
            const fileNameDiv = document.getElementById('file-name');
            if (input.files.length > 0) {
                fileNameDiv.innerText = "Selected: " + input.files[0].name;
            }
        }

        document.getElementById('uploadForm').addEventListener('submit', async function(e) {
            e.preventDefault();
            const fileInput = document.getElementById('pdfFile');
            if (fileInput.files.length === 0) return;

            const submitBtn = document.getElementById('submitBtn');
            const spinner = document.getElementById('spinner');

            submitBtn.disabled = true;
            submitBtn.innerText = "Building Pro Dashboard...";
            spinner.style.display = "block";

            const formData = new FormData();
            formData.append("file", fileInput.files[0]);

            try {
                const response = await fetch("/convert-pdf/", { method: "POST", body: formData });
                if (!response.ok) throw new Error("Conversion failed!");

                const blob = await response.blob();
                const downloadUrl = window.URL.createObjectURL(blob);
                const a = document.createElement("a");
                a.href = downloadUrl;
                a.download = fileInput.files[0].name.replace(".pdf", "_ProDashboard.xlsx");
                document.body.appendChild(a);
                a.click();
                a.remove();
            } catch (err) {
                alert("Error extracting PDF data.");
            } finally {
                submitBtn.disabled = false;
                submitBtn.innerText = "⚡ Generate Designer Excel";
                spinner.style.display = "none";
            }
        });
    </script>
</body>
</html>
"""

@app.get("/", response_class=HTMLResponse)
def serve_ui():
    return HTML_PAGE

# -------------------------------------------------------------
# ⚙️ EXTRACTION LOGIC
# -------------------------------------------------------------
def extract_clean_number(val_str: str) -> float:
    if not val_str:
        return 0.0
    main_part = val_str.split('(')[0]
    clean_str = re.sub(r"[^\d.]", "", main_part)
    try:
        return float(clean_str)
    except:
        return 0.0

def parse_exact_pdf_page(page, page_num: int) -> dict:
    row_data = {
        "Page_No": page_num,
        "Customer_Name": "",
        "Product": "",
        "Quantity": 0,
        "Unit_Price": 0.0,
        "Base_Amount": 0.0,
        "Discount": "",
        "Taxable_Amount": 0.0,
        "GST": 0.0,
        "Net_Total": 0.0,
        "Description": ""
    }

    tables = page.extract_tables()
    
    if tables:
        for table in tables:
            for row in table:
                if len(row) >= 2 and row[0] and row[1]:
                    field = str(row[0]).replace("\n", " ").strip().lower()
                    val = str(row[1]).replace("\n", " ").strip()
                    
                    if "customer" in field: row_data["Customer_Name"] = val
                    elif "product" in field: row_data["Product"] = val
                    elif "quantity" in field:
                        try: row_data["Quantity"] = int(re.sub(r"[^\d]", "", val))
                        except: pass
                    elif "unit price" in field: row_data["Unit_Price"] = extract_clean_number(val)
                    elif "base amount" in field: row_data["Base_Amount"] = extract_clean_number(val)
                    elif "discount" in field: row_data["Discount"] = val
                    elif "taxable" in field: row_data["Taxable_Amount"] = extract_clean_number(val)
                    elif "gst" in field: row_data["GST"] = extract_clean_number(val)
                    elif "money received" in field or "net total" in field: row_data["Net_Total"] = extract_clean_number(val)
                    elif "description" in field: row_data["Description"] = val
    else:
        text = page.extract_text() or ""
        lines = text.split("\n")
        for line in lines:
            if "|" in line or ":" in line:
                parts = re.split(r"[:|]", line, 1)
                if len(parts) == 2:
                    field = parts[0].strip().lower()
                    val = parts[1].strip()
                    
                    if "customer" in field: row_data["Customer_Name"] = val
                    elif "product" in field: row_data["Product"] = val
                    elif "quantity" in field:
                        try: row_data["Quantity"] = int(re.sub(r"[^\d]", "", val))
                        except: pass
                    elif "unit price" in field: row_data["Unit_Price"] = extract_clean_number(val)
                    elif "description" in field: row_data["Description"] = val

    return row_data

# -------------------------------------------------------------
# 🚀 FASTAPI DESIGNER EXCEL CONVERT ROUTE
# -------------------------------------------------------------
@app.post("/convert-pdf/")
async def convert_pdf(file: UploadFile = File(...)):
    if not file.filename.endswith('.pdf'):
        raise HTTPException(status_code=400, detail="Only PDF files allowed.")
    
    try:
        contents = await file.read()
        pdf_file = io.BytesIO(contents)
        pages_data = []
        
        with pdfplumber.open(pdf_file) as pdf:
            for i, page in enumerate(pdf.pages, start=1):
                pages_data.append(parse_exact_pdf_page(page, i))
        
        wb = openpyxl.Workbook()
        
        # -------------------------------------------------------------
        # 👑 STYLING PALETTE
        # -------------------------------------------------------------
        font_family = "Segoe UI"
        
        navy_header_fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
        card_label_fill = PatternFill(start_color="F0F4F8", end_color="F0F4F8", fill_type="solid")
        accent_green_fill = PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid")
        
        white_banner_font = Font(name=font_family, size=13, bold=True, color="FFFFFF")
        label_bold_font = Font(name=font_family, size=10, bold=True, color="333333")
        payable_bold_font = Font(name=font_family, size=12, bold=True, color="1E7E34")

        thin_side = Side(border_style="thin", color="D1D5DB")
        thick_green_side = Side(border_style="medium", color="28A745")
        
        card_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        payable_border = Border(left=thick_green_side, right=thick_green_side, top=thick_green_side, bottom=thick_green_side)

        # -------------------------------------------------------------
        # 1. PAGEVIEWER TAB
        # -------------------------------------------------------------
        ws_view = wb.active
        ws_view.title = "PageViewer"
        ws_view.views.sheetView[0].showGridLines = True

        # Header Banner
        ws_view.merge_cells("B2:D2")
        ws_view["B2"] = "INVOICE DATA EXECUTIVE DASHBOARD"
        ws_view["B2"].font = white_banner_font
        ws_view["B2"].fill = navy_header_fill
        ws_view["B2"].alignment = Alignment(horizontal="center", vertical="center")
        ws_view.row_dimensions[2].height = 35

        # Dropdown Cell Setup with High Contrast
        ws_view["B4"] = "Select Page No:"
        ws_view["B4"].font = label_bold_font
        ws_view["B4"].alignment = Alignment(horizontal="right", vertical="center")
        
        ws_view["C4"] = 1
        ws_view["C4"].font = Font(name=font_family, size=11, bold=True, color="007BFF")
        ws_view["C4"].fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        ws_view["C4"].alignment = Alignment(horizontal="center", vertical="center")
        ws_view["C4"].border = Border(
            left=Side(style="medium", color="007BFF"),
            right=Side(style="medium", color="007BFF"),
            top=Side(style="medium", color="007BFF"),
            bottom=Side(style="medium", color="007BFF")
        )
        ws_view.row_dimensions[4].height = 25

        # Forced Dropdown Binding
        page_numbers_list = ",".join(str(i) for i in range(1, len(pages_data) + 1))
        dv = DataValidation(type="list", formula1=f'"{page_numbers_list}"', allow_blank=False, showDropDown=True)
        dv.error = 'Kripya list se valid Page Number select karein!'
        dv.errorTitle = 'Invalid Page'
        ws_view.add_data_validation(dv)
        dv.add(ws_view["C4"])

        # Form Rows
        calc_rows = [
            ("Customer Name", '=XLOOKUP(C4, DataSheet!A:A, DataSheet!B:B, "N/A")', "@"),
            ("Product Name", '=XLOOKUP(C4, DataSheet!A:A, DataSheet!C:C, "N/A")', "@"),
            ("Quantity", '=XLOOKUP(C4, DataSheet!A:A, DataSheet!D:D, 0)', "#,##0"),
            ("Unit Price", '=XLOOKUP(C4, DataSheet!A:A, DataSheet!E:E, 0)', "₹#,##0.00"),
            ("Base Amount", '=XLOOKUP(C4, DataSheet!A:A, DataSheet!F:F, 0)', "₹#,##0.00"),
            ("Discount", '=XLOOKUP(C4, DataSheet!A:A, DataSheet!G:G, "0")', "@"),
            ("Taxable Amount", '=XLOOKUP(C4, DataSheet!A:A, DataSheet!H:H, 0)', "₹#,##0.00"),
            ("GST Amount", '=XLOOKUP(C4, DataSheet!A:A, DataSheet!I:I, 0)', "₹#,##0.00"),
            ("FINAL PAYABLE", '=XLOOKUP(C4, DataSheet!A:A, DataSheet!J:J, 0)', "₹#,##0.00"),
            ("Description", '=XLOOKUP(C4, DataSheet!A:A, DataSheet!K:K, "")', "@")
        ]

        row_idx = 6
        for label, formula, num_format in calc_rows:
            ws_view.row_dimensions[row_idx].height = 22
            
            ws_view[f"B{row_idx}"] = label
            ws_view[f"B{row_idx}"].font = label_bold_font
            ws_view[f"B{row_idx}"].fill = card_label_fill
            ws_view[f"B{row_idx}"].border = card_border
            ws_view[f"B{row_idx}"].alignment = Alignment(horizontal="left", vertical="center", indent=1)

            ws_view.merge_cells(f"C{row_idx}:D{row_idx}")
            cell = ws_view[f"C{row_idx}"]
            cell.value = formula
            cell.number_format = num_format
            cell.border = card_border
            cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

            if label == "FINAL PAYABLE":
                ws_view[f"B{row_idx}"].fill = accent_green_fill
                ws_view[f"B{row_idx}"].font = Font(name=font_family, size=11, bold=True, color="1E7E34")
                cell.fill = accent_green_fill
                cell.font = payable_bold_font
                ws_view[f"B{row_idx}"].border = payable_border
                cell.border = payable_border

            row_idx += 1

        ws_view.column_dimensions['A'].width = 3
        ws_view.column_dimensions['B'].width = 24
        ws_view.column_dimensions['C'].width = 20
        ws_view.column_dimensions['D'].width = 20

        # -------------------------------------------------------------
        # 2. DATASHEET TAB
        # -------------------------------------------------------------
        ws_data = wb.create_sheet(title="DataSheet")
        ws_data.views.sheetView[0].showGridLines = True

        headers = [
            "Page_No", "Customer_Name", "Product", "Quantity", "Unit_Price",
            "Base_Amount", "Discount", "Taxable_Amount", "GST_Amount",
            "Money_Received", "Description"
        ]
        ws_data.append(headers)

        ws_data.row_dimensions[1].height = 26
        for col_num, header in enumerate(headers, 1):
            cell = ws_data.cell(row=1, column=col_num)
            cell.font = Font(name=font_family, size=10, bold=True, color="FFFFFF")
            cell.fill = navy_header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for idx, item in enumerate(pages_data, start=2):
            ws_data.append([
                item["Page_No"], item["Customer_Name"], item["Product"],
                item["Quantity"], item["Unit_Price"], item["Base_Amount"],
                item["Discount"], item["Taxable_Amount"], item["GST"],
                item["Net_Total"], item["Description"]
            ])
            
            ws_data.row_dimensions[idx].height = 20
            ws_data[f"E{idx}"].number_format = "₹#,##0.00"
            ws_data[f"F{idx}"].number_format = "₹#,##0.00"
            ws_data[f"H{idx}"].number_format = "₹#,##0.00"
            ws_data[f"I{idx}"].number_format = "₹#,##0.00"
            ws_data[f"J{idx}"].number_format = "₹#,##0.00"

        for col in ws_data.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws_data.column_dimensions[col_letter].width = max(max_len + 3, 12)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={'Content-Disposition': f'attachment; filename="{file.filename.replace(".pdf", "_ProDashboard.xlsx")}"'}
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
